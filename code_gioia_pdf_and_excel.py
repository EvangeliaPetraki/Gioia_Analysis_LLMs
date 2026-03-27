"""
WP5 Gioia First-Order Concepts Extractor
------------------------------------------------
- Reads PDFs from an input folder
- Chunks text
- Calls LLM per chunk through API (STRICT JSON for Gioia 1st-order concepts)
- Writes ONE PDF PER INPUT PDF into an output folder
- Local merge across chunks (dedupe concepts)
"""

import os
import re
import json
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import requests
import pdfplumber
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
import random
from reportlab.pdfgen import canvas
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from analysis.env_config import get_bool, get_float, get_int, get_list, get_path, get_str


# ----------------------------
# Config
# ----------------------------
MAX_RETRIES = get_int("GIOIA_FIRST_ORDER_MAX_RETRIES", 8)
BASE_BACKOFF_S = get_float("GIOIA_FIRST_ORDER_BASE_BACKOFF_S", 2.0)
MAX_BACKOFF_S = get_float("GIOIA_FIRST_ORDER_MAX_BACKOFF_S", 60.0)

CHARS_PER_CHUNK = get_int("CHARS_PER_CHUNK", 4000)
SLEEP_BETWEEN_CALLS = get_float("SLEEP_BETWEEN_CALLS", 0.5)

CHUTES_API_KEY = get_str("CHUTES_API_KEY", "")
CHUTES_BASE_URL = get_str("CHUTES_BASE_URL", "https://llm.chutes.ai/v1")
CHUTES_MODEL = get_str(
    "GIOIA_FIRST_ORDER_MODEL",
    get_str("CHUTES_MODEL", "deepseek-ai/DeepSeek-V3-0324"),
)

# Prefer hot / high-availability models for batch jobs
FALLBACK_MODELS = get_list(
    "GIOIA_FIRST_ORDER_FALLBACK_MODELS",
    [
        CHUTES_MODEL,
        "Qwen/Qwen3-32B",
        "mistral/Mistral-Nemo-Instruct-2407",
        "deepseek/DeepSeek-V3-0324",
    ],
)

# Controls PDF verbosity:
INCLUDE_CHUNK_OUTPUTS = get_bool("GIOIA_FIRST_ORDER_INCLUDE_CHUNK_OUTPUTS", False)
INCLUDE_MERGED_RECORD = get_bool("GIOIA_FIRST_ORDER_INCLUDE_MERGED_RECORD", True)

EXCEL_FILENAME = get_str("GIOIA_FIRST_ORDER_EXCEL_FILENAME", "GIOIA_FIRST_ORDER_CONCEPTS.xlsx")
EXCEL_HEADERS = get_list(
    "GIOIA_FIRST_ORDER_EXCEL_HEADERS",
    ["country", "file_name", "first_order_concept", "concept_type", "quote", "doc_id"],
)


CAPACITY_STATUS = {429, 503}

def _parse_http_status_from_error(e: Exception) -> Optional[int]:
    m = re.search(r"Chutes HTTP (\d{3})", str(e))
    return int(m.group(1)) if m else None

def _is_capacity_error(e: Exception) -> bool:
    msg = (str(e) or "").lower()
    status = _parse_http_status_from_error(e)
    if status in CAPACITY_STATUS:
        # match common Chutes capacity messages
        return (
            "maximum capacity" in msg
            or "try again later" in msg
            or "no instances available" in msg
            or "infrastructure is at maximum capacity" in msg
        )
    return False

def _exp_backoff_with_jitter(attempt_idx: int, base: float = BASE_BACKOFF_S, cap: float = MAX_BACKOFF_S) -> float:
    # attempt_idx is 0,1,2...
    delay = min(cap, base * (2 ** attempt_idx))
    return delay * (0.7 + 0.6 * random.random())  # jitter 0.7–1.3x


# ----------------------------
# Prompts (Gioia Stage 1)
# ----------------------------
SYSTEM_GIOIA_FIRST_ORDER = """You are assisting a Horizon Europe research project (WP5 – European labour market resilience management). You are acting as a structured qualitative coding assistant. The researcher retains interpretive authority.

We are applying the Gioia methodology (Gioia, Corley & Hamilton, 2013).
In THIS step, you must extract FIRST-ORDER CONCEPTS from the provided policy text.

Definition (operational):
First-order concepts are:
- Close to the language of the policy document
- Descriptive rather than interpretive
- Reflecting how the policy frames problems, objectives, instruments, target groups, governance, or transition challenges
- Not abstract theoretical categories
- Not grouped into broader themes
- Not evaluated or critiqued

Important distinctions:
- Do NOT create second-order themes.
- Do NOT summarise the whole document.
-Each concept must represent one distinct intervention logic or policy-relevant idea. Do not artificially split a single intervention into multiple concepts.
- Each concept must represent one distinct idea explicitly stated in the text.
- Ignore section headings, figure titles, and numbered section labels.

Each concept must:
- Have a short label (3–10 words)
- Be supported by one short verbatim quote from the text (<=25 words)
- Stay faithful to the document’s meaning


RELEVANCE FILTER (CRITICAL):

Extract first-order concepts ONLY if they relate to:

- Labour market impacts, employment, skills, reskilling, upskilling
- Governance mechanisms for green or digital transition
- Policy instruments, funding tools, regulatory measures
- Social inclusion, inequality, vulnerable groups
- Resilience, adaptability, shock absorption, transition management
- Regional or sectoral transition challenges
- Institutional or regulatory change related to labour markets

DO NOT extract:
- Literature titles or bibliographic references
- Section headings or numbering
- Methodological descriptions of data analysis
- Database descriptions
- Annex information
- Website/contact details
- Copyright or legal disclaimers
- Generic background information without policy relevance

Granularity rule:

If multiple statements refer to the same intervention logic, governance mechanism, or policy instrument, consolidate them into ONE first-order concept.

Do not create multiple concepts that only differ in wording but refer to the same type of action.

Prefer slightly broader first-order labels when several sentences describe the same intervention.

If content is purely descriptive of academic analysis methods, ignore it.
If content is purely bibliographic, ignore it.
If content is structural (headings, annex titles), ignore it.

If the text contains repeated ideas, only include meaningfully distinct concepts.

PER-CHUNK DISCIPLINE RULES:

You are analysing ONE CHUNK of a larger policy document.

Important:
- Do NOT attempt to infer content outside this chunk.
- Extract only concepts clearly supported within this chunk.
- Do NOT try to anticipate later sections of the document.
- Do NOT aim for exhaustiveness within this chunk.

Concept density control:
- Prefer 5–12 high-quality, non-redundant concepts per chunk.
- Only extract concepts that are clearly central to policy content.
- If the chunk contains limited policy-relevant material, return fewer concepts.
- If no relevant policy content appears, return an empty list.

Cross-chunk consistency:
- If similar ideas appear multiple times within this chunk, consolidate them into one concept.
- Avoid creating slightly reworded duplicates.

ADDITIONAL METADATA EXTRACTION:

Before listing first-order concepts, identify and clearly state:

1. Governance level:
   - Supranational (e.g., EU, OECD, UN)
   - National
   - Regional
   - Local
   - Sectoral / Industry-specific

2. Country (or countries if supranational)

3. Region (if applicable)

4. Policy type (choose the closest category):
   - Regulation / Law
   - Strategy / Agenda
   - Funding Instrument
   - Framework / Governance Mechanism
   - Partnership / Coalition
   - Programme / Initiative
   - Research / Knowledge Instrument

If information is not explicitly stated in the document, write: "Not explicitly specified in document".

OUTPUT FORMAT:

Document Metadata:
- Governance Level:
- Country:
- Region:
- Policy Type:
- Year (if mentioned):

OUTPUT SCHEMA:
{
  "chunk_id": int,
  "doc_id": string|null,
  "source_pages": {"start": int|null, "end": int|null},
  "document_metadata": {
    "document_type": "policy_instrument"|"policy_report"|"research_paper"|"discussion_paper"|"technical_study"|"case_study"|"other"|"unknown",
    "governance_level": "supranational"|"national"|"regional"|"local"|"sectoral"|"unknown",
    "country": string|null,
    "region": string|null,
    "issuing_body": string|null,
    "year": int|null
  },
  "first_order_concepts": [
    {
      "label": string,
      "type": "objective"|"measure"|"instrument"|"target_group"|"governance"|"funding"|"problem_framing"|"resilience_framing"|"twin_transition_green"|"twin_transition_digital"|"social_inclusion"|"monitoring_evaluation"|"other",
      "quote": string
    }
  ]
}

Hard constraints:
- label max 80 chars
- quote max 240 chars and <=25 words (approx.)
- first_order_concepts max 30
- Output must start with '{' and end with '}'.
"""

USER_GIOIA_TEMPLATE = """chunk_id: {chunk_id}
doc_id: {doc_id}
pages: {page_start}-{page_end}
text:
<<<
{chunk_text}
>>>
"""

SYSTEM_JSON_REPAIR = r"""You are a strict JSON repair tool.
You will receive text that is intended to contain ONE JSON object but may be invalid, truncated, or wrapped.
Return a single VALID JSON object only.
Rules:
- Output must start with '{' and end with '}'.
- No markdown, no commentary.
- Fix quoting/escaping issues as needed.
- IMPORTANT: Do NOT emit \\uXXXX escapes. Use actual UTF-8 characters instead.
- If multiple JSON objects appear, return the most complete one.
If repair is impossible, return: {"error":"unrepairable_json"}.
"""


# ----------------------------
# JSON parsing helpers
# ----------------------------

_HEX4 = r"[0-9a-fA-F]{4}"
_HEX8 = r"[0-9a-fA-F]{8}"
_HEX2 = r"[0-9a-fA-F]{2}"

def sanitize_invalid_json_escapes(s: str) -> str:
    # """
    # Make common LLM JSON breakages parseable by json.loads by fixing *invalid* escape sequences.
    # Key fix: turn '\\u' that is NOT followed by 4 hex digits into '\\\\u' (literal backslash-u).
    # Also covers \UXXXXXXXX and \xXX.
    # """
    if not s:
        return s

    # Remove BOM if present
    s = s.lstrip("\ufeff")

    # Fix invalid unicode escape starters
    s = re.sub(rf"\\u(?!{_HEX4})", r"\\\\u", s)
    s = re.sub(rf"\\U(?!{_HEX8})", r"\\\\U", s)
    s = re.sub(rf"\\x(?!{_HEX2})", r"\\\\x", s)

    # If the model produced a stray backslash at end (rare but fatal), escape it
    s = re.sub(r"\\\s*$", r"\\\\", s)

    return s


def _extract_json_obj(text: str) -> Dict[str, Any]:
    text = (text or "").strip()

    # 1) First try direct parse
    try:
        return json.loads(text)
    except Exception:
        pass

    # 2) Try sanitize + parse (fix invalid \u escapes etc.)
    try:
        return json.loads(sanitize_invalid_json_escapes(text))
    except Exception:
        pass

    # 3) Try extracting the last {...} block, then parse (with sanitize fallback)
    m = re.search(r"(\{.*\})\s*$", text, flags=re.S)
    if m:
        candidate = m.group(1)
        try:
            return json.loads(candidate)
        except Exception:
            return json.loads(sanitize_invalid_json_escapes(candidate))

    raise ValueError("No valid JSON object found in model output.")


def safe_parse_json(text: str) -> Dict[str, Any]:
    return _extract_json_obj(text)


# ----------------------------
# Chutes LLM call
# ----------------------------
def llm_raw(
    system_prompt: str,
    user_prompt: str,
    *,
    model: Optional[str] = None,
    temperature: float = 0.0,
    max_tokens: int = 3500
) -> str:
    if not CHUTES_API_KEY:
        raise RuntimeError("Set CHUTES_API_KEY in your environment.")

    url = f"{CHUTES_BASE_URL.rstrip('/')}/chat/completions"
    payload = {
        "model": model or CHUTES_MODEL,
        "temperature": temperature,
        "max_tokens": max_tokens,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    }
    headers = {
        "Authorization": f"Bearer {CHUTES_API_KEY}",
        "Content-Type": "application/json",
    }

    r = requests.post(url, headers=headers, json=payload, timeout=120)

    if not r.ok:
        with open("last_http_error.txt", "w", encoding="utf-8") as f:
            f.write(f"STATUS: {r.status_code}\n")
            f.write(r.text)
        raise RuntimeError(f"Chutes HTTP {r.status_code}: {r.text[:2000]}")

    data = r.json()
    return data["choices"][0]["message"]["content"]

def llm_json(
    system_prompt: str,
    user_prompt: str,
    *,
    model: Optional[str] = None,
    temperature: float = 0.0,
    max_tokens: int = 3500,
    retries: int = MAX_RETRIES,
) -> Dict[str, Any]:
    last_err: Optional[Exception] = None

    # rotate models on capacity errors
    models_to_try = [model] if model else []
    models_to_try += [m for m in FALLBACK_MODELS if m and m not in models_to_try]

    attempt = 0
    while attempt < retries:
        chosen_model = models_to_try[attempt % len(models_to_try)]
        text = ""

        try:
            text = llm_raw(system_prompt, user_prompt, model=chosen_model,
                           temperature=temperature, max_tokens=max_tokens)

            try:
                return safe_parse_json(text)
            except Exception:
                with open("last_bad_output.txt", "w", encoding="utf-8") as f:
                    f.write(text)

                repaired = llm_raw(
                    SYSTEM_JSON_REPAIR,
                    f"TEXT:\n<<<\n{text}\n>>>",
                    model=chosen_model,
                    temperature=0.0,
                    max_tokens=max_tokens
                )
                with open("last_repaired_output.txt", "w", encoding="utf-8") as f:
                    f.write(repaired)
                return safe_parse_json(repaired)

        except Exception as e:
            last_err = e
            if text:
                with open("last_exception_output.txt", "w", encoding="utf-8") as f:
                    f.write(text)

            # Capacity-aware handling
            if _is_capacity_error(e):
                delay = _exp_backoff_with_jitter(attempt)
                print(f"[WARN] Capacity error on model={chosen_model}. Backing off {delay:.1f}s (attempt {attempt+1}/{retries})")
                time.sleep(delay)
            else:
                # non-capacity errors: small backoff
                delay = min(5.0, 0.8 * (attempt + 1))
                print(f"[WARN] Non-capacity error on model={chosen_model}. Sleeping {delay:.1f}s (attempt {attempt+1}/{retries})")
                time.sleep(delay)

            attempt += 1

    raise RuntimeError(f"LLM call failed after {retries} attempts: {last_err}")

# ----------------------------
# PDF text extraction + chunking
# ----------------------------
def clean_text(t: str) -> str:
    t = re.sub(r"\s+\n", "\n", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    t = re.sub(r"[ \t]{2,}", " ", t)
    return t.strip()


def extract_pdf_pages(pdf_path: str) -> List[Tuple[int, str]]:
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            txt = page.extract_text() or ""
            pages.append((i, txt))
    return pages


def chunk_pages(pages: List[Tuple[int, str]], max_chars: int = CHARS_PER_CHUNK):
    buf: List[Tuple[int, str]] = []
    buf_len = 0
    start_page: Optional[int] = None

    for pno, txt in pages:
        txt = clean_text(txt or "")
        if not txt:
            continue

        if start_page is None:
            start_page = pno

        if buf_len + len(txt) > max_chars and buf:
            yield {
                "page_start": start_page,
                "page_end": buf[-1][0],
                "text": "\n\n".join([f"[PAGE {pno}]\n{ptxt}" for (pno, ptxt) in buf]),
            }
            buf, buf_len = [], 0
            start_page = pno

        buf.append((pno, txt))
        buf_len += len(txt)

    if buf:
        yield {
            "page_start": start_page,
            "page_end": buf[-1][0],
            "text": "\n\n".join([f"[PAGE {pno}]\n{ptxt}" for (pno, ptxt) in buf]),
        }


# ----------------------------
# Local merge (concept-level)
# ----------------------------
def merge_chunk_concepts_locally(chunk_jsons: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Merge chunk outputs into one doc-level object:
    - Merge page ranges (min start, max end)
    - Dedupe concepts by normalized (label + type)
    - Keep first quote seen (you can later decide to keep multiple)
    """
    def norm(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").strip().lower())

    doc_id = None
    min_page = None
    max_page = None

    merged_map: Dict[str, Dict[str, Any]] = {}

    for obj in chunk_jsons:
        if not isinstance(obj, dict):
            continue

        if not doc_id and obj.get("doc_id"):
            doc_id = obj.get("doc_id")

        sp = obj.get("source_pages") or {}
        if isinstance(sp, dict):
            s = sp.get("start")
            e = sp.get("end")
            if isinstance(s, int):
                min_page = s if (min_page is None or s < min_page) else min_page
            if isinstance(e, int):
                max_page = e if (max_page is None or e > max_page) else max_page

        concepts = obj.get("first_order_concepts") or []
        if not isinstance(concepts, list):
            continue

        for c in concepts:
            if not isinstance(c, dict):
                continue
            label = c.get("label") or ""
            ctype = c.get("type") or "other"
            if not str(label).strip():
                continue
            key = f"{norm(label)}||{norm(ctype)}"
            if key not in merged_map:
                merged_map[key] = {
                    "label": str(label)[:80],
                    "type": ctype,
                    "quote": (c.get("quote") or "")[:240],
                }

    # stable ordering: type then label
    merged_list = sorted(
        merged_map.values(),
        key=lambda x: (str(x.get("type") or ""), str(x.get("label") or "").lower())
    )

    return {
        "doc_id": doc_id,
        "source_pages": {"start": min_page, "end": max_page},
        "first_order_concepts": merged_list,
    }


# ----------------------------
# PDF writing (one output PDF per input PDF)
# ----------------------------
def _wrap_lines(text: str, max_chars: int = 110):
    text = (text or "").strip()
    if not text:
        return []
    text = re.sub(r"\s+", " ", text)
    return [text[i:i + max_chars] for i in range(0, len(text), max_chars)]


def write_concepts_pdf(
    source_pdf_path: Path,
    output_dir: Path,
    doc_id: str,
    merged: Optional[Dict[str, Any]],
    chunk_outputs: List[Dict[str, Any]],
    include_merged: bool = True,
    include_chunks: bool = True,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    out_pdf = output_dir / f"{source_pdf_path.stem}_GIOIA_FIRST_ORDER.pdf"

    c = canvas.Canvas(str(out_pdf), pagesize=A4)
    width, height = A4
    x = 2 * cm
    y = height - 2 * cm

    # Header
    c.setFont("Helvetica-Bold", 14)
    c.drawString(x, y, "WP5 Gioia — First-Order Concepts (LLM-assisted)")
    y -= 18

    c.setFont("Helvetica", 10)
    intro = (
        "This report lists first-order concepts extracted using a Gioia-method stage-1 procedure. "
        "Concepts stay close to the document language and are supported by short verbatim quotes. "
        "No interpretation, evaluation, or second-order themes are produced here."
    )
    for ln in _wrap_lines(intro, max_chars=110):
        c.drawString(x, y, ln)
        y -= 12

    y -= 4
    c.setFont("Helvetica", 10)
    c.drawString(x, y, f"Document ID: {doc_id}")
    y -= 14
    c.drawString(x, y, f"Source file: {source_pdf_path.name}")
    y -= 18

        # Document metadata (from first chunk if available)
    meta = None
    if chunk_outputs and isinstance(chunk_outputs[0].get("extracted"), dict):
        meta = chunk_outputs[0]["extracted"].get("document_metadata")

    if isinstance(meta, dict):
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x, y, "Document metadata (from model output):")
        y -= 14
        c.setFont("Helvetica", 10)

        lines = [
            f"Document type: {meta.get('document_type')}",
            f"Governance level: {meta.get('governance_level')}",
            f"Country: {meta.get('country') or '—'}",
            f"Region: {meta.get('region') or '—'}",
            f"Issuing body: {meta.get('issuing_body') or '—'}",
            f"Year: {meta.get('year') or '—'}",
        ]
        for ln in lines:
            if y < 4 * cm:
                c.showPage()
                y = height - 2 * cm
                c.setFont("Helvetica", 10)
            c.drawString(x, y, ln[:120])
            y -= 12
        y -= 6


    # Merged concepts
    if include_merged and merged and isinstance(merged.get("first_order_concepts"), list):
        concepts = merged.get("first_order_concepts") or []
        if y < 4 * cm:
            c.showPage()
            y = height - 2 * cm

        c.setFont("Helvetica-Bold", 12)
        c.drawString(x, y, "Document-Level First-Order Concepts (merged across chunks)")
        y -= 16

        c.setFont("Helvetica", 10)
        if not concepts:
            c.drawString(x, y, "No first-order concepts were extracted from this document.")
            y -= 14
        else:
            for idx, cc in enumerate(concepts, start=1):
                if y < 3.5 * cm:
                    c.showPage()
                    y = height - 2 * cm

                label = (cc.get("label") or "").strip()
                ctype = (cc.get("type") or "other").strip()
                quote = (cc.get("quote") or "").strip()

                c.setFont("Helvetica-Bold", 10)
                c.drawString(x, y, f"{idx}. {label}  [{ctype}]")
                y -= 12

                if quote:
                    c.setFont("Helvetica-Oblique", 9)
                    for ln in _wrap_lines(f"Evidence: “{quote}”", max_chars=105):
                        c.drawString(x + 12, y, ln)
                        y -= 11
                y -= 6

    # Chunk appendix
    if include_chunks:
        c.showPage()
        y = height - 2 * cm
        c.setFont("Helvetica-Bold", 12)
        c.drawString(x, y, "Appendix — Chunk-Level Output (for validation)")
        y -= 16

        c.setFont("Helvetica", 10)
        c.drawString(x, y, "This section prints the raw JSON outputs per chunk.")
        y -= 14

        for i, ch in enumerate(chunk_outputs, start=1):
            if y < 4 * cm:
                c.showPage()
                y = height - 2 * cm

            c.setFont("Helvetica-Bold", 10)
            c.drawString(x, y, f"Chunk {i} | pages {ch['page_start']}-{ch['page_end']}")
            y -= 14

            extracted = ch.get("extracted") or {}
            dump = json.dumps(extracted, ensure_ascii=False, indent=2)

            c.setFont("Courier", 7)
            for ln in dump.splitlines():
                if y < 2.5 * cm:
                    c.showPage()
                    y = height - 2 * cm
                    c.setFont("Courier", 7)
                # truncate very long lines for PDF safety
                ln = ln[:160]
                c.drawString(x, y, ln)
                y -= 9

            y -= 10

    c.save()
    return out_pdf


# ----------------------------
# Excel writing (one common workbook for all input PDFs)
# ----------------------------
def infer_country_from_pdf_path(source_pdf_path: Path) -> str:
    parts_lower = [p.lower() for p in source_pdf_path.parts]
    if "input" in parts_lower:
        idx = parts_lower.index("input")
        if idx > 0:
            return source_pdf_path.parts[idx - 1]
    return source_pdf_path.parent.name


def init_concepts_workbook(output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    excel_path = output_dir / EXCEL_FILENAME

    if not excel_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "first_order_concepts"
        ws.append(EXCEL_HEADERS)
        for i, header in enumerate(EXCEL_HEADERS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(18, len(header) + 2)
        wb.save(excel_path)

    return excel_path


def _existing_doc_keys(excel_path: Path) -> set:
    if not excel_path.exists():
        return set()

    wb = load_workbook(excel_path)
    ws = wb.active
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        country = row[0] if len(row) > 0 else None
        file_name = row[1] if len(row) > 1 else None
        if country or file_name:
            keys.add((str(country or ""), str(file_name or "")))
    wb.close()
    return keys


def write_concepts_excel(
    excel_path: Path,
    source_pdf_path: Path,
    doc_id: str,
    merged: Optional[Dict[str, Any]],
) -> Path:
    country = infer_country_from_pdf_path(source_pdf_path)
    file_name = source_pdf_path.name

    wb = load_workbook(excel_path)
    ws = wb.active

    concepts = []
    if merged and isinstance(merged.get("first_order_concepts"), list):
        concepts = merged.get("first_order_concepts") or []

    if concepts:
        for cc in concepts:
            if not isinstance(cc, dict):
                continue
            ws.append([
                country,
                file_name,
                (cc.get("label") or "").strip(),
                (cc.get("type") or "other").strip(),
                (cc.get("quote") or "").strip(),
                doc_id,
            ])
    else:
        ws.append([country, file_name, "", "", "", doc_id])

    wb.save(excel_path)
    wb.close()
    return excel_path


# ----------------------------
# Main processing: one PDF -> one output PDF + append to common Excel
# ----------------------------
def process_one_pdf(pdf_path: str, output_folder: str, excel_path: Path, existing_doc_keys: Optional[set] = None) -> Path:
    source_pdf = Path(pdf_path)
    doc_id = source_pdf.stem
    out_dir = Path(output_folder)

    out_pdf = out_dir / f"{source_pdf.stem}_GIOIA_FIRST_ORDER.pdf"
    doc_key = (infer_country_from_pdf_path(source_pdf), source_pdf.name)

    if out_pdf.exists() and existing_doc_keys is not None and doc_key in existing_doc_keys:
        print(f"Skipping (PDF and Excel already exist): {source_pdf.name}")
        return out_pdf

    pages = extract_pdf_pages(str(source_pdf))
    chunks = list(chunk_pages(pages))

    chunk_outputs: List[Dict[str, Any]] = []
    for i, ch in enumerate(chunks, start=1):
        user_prompt = USER_GIOIA_TEMPLATE.format(
            chunk_id=i,
            doc_id=doc_id,
            page_start=ch["page_start"],
            page_end=ch["page_end"],
            chunk_text=ch["text"],
        )

        extracted = llm_json(SYSTEM_GIOIA_FIRST_ORDER, user_prompt, max_tokens=3500)

        # Ensure minimal required keys exist (defensive)
        if "chunk_id" not in extracted:
            extracted["chunk_id"] = i
        if "doc_id" not in extracted:
            extracted["doc_id"] = doc_id
        if "source_pages" not in extracted:
            extracted["source_pages"] = {"start": ch["page_start"], "end": ch["page_end"]}
        if "first_order_concepts" not in extracted or not isinstance(extracted["first_order_concepts"], list):
            extracted["first_order_concepts"] = []
        if "document_metadata" not in extracted or not isinstance(extracted["document_metadata"], dict):
            extracted["document_metadata"] = {
                "document_type": "unknown",
                "governance_level": "unknown",
                "country": None,
                "region": None,
                "issuing_body": None,
                "year": None,
            }

        chunk_outputs.append({
            "page_start": ch["page_start"],
            "page_end": ch["page_end"],
            "extracted": extracted,
        })
        time.sleep(SLEEP_BETWEEN_CALLS)

    extracted_list = [x["extracted"] for x in chunk_outputs]
    merged = merge_chunk_concepts_locally(extracted_list) if INCLUDE_MERGED_RECORD else None

    out_pdf = write_concepts_pdf(
        source_pdf_path=source_pdf,
        output_dir=out_dir,
        doc_id=doc_id,
        merged=merged,
        chunk_outputs=chunk_outputs,
        include_merged=INCLUDE_MERGED_RECORD,
        include_chunks=INCLUDE_CHUNK_OUTPUTS,
    )

    if existing_doc_keys is None or doc_key not in existing_doc_keys:
        out_excel = write_concepts_excel(
            excel_path=excel_path,
            source_pdf_path=source_pdf,
            doc_id=doc_id,
            merged=merged,
        )
        print(f"Saved concepts to Excel: {out_excel}")
        if existing_doc_keys is not None:
            existing_doc_keys.add(doc_key)
    else:
        print(f"Skipping Excel append (already in workbook): {source_pdf.name}")

    print(f"Saved concepts PDF: {out_pdf}")
    return out_pdf


def process_folder(input_folder: str, output_folder: str):
    input_folder = Path(input_folder)
    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)

    excel_path = init_concepts_workbook(output_folder)
    existing_doc_keys = _existing_doc_keys(excel_path)

    pdfs = sorted([p for p in input_folder.iterdir() if p.suffix.lower() == ".pdf"])
    print(f"Found {len(pdfs)} PDFs in {input_folder}")
    print(f"Common Excel file: {excel_path.name}")

    for p in pdfs:
        print(f"\nProcessing: {p.name}")
        try:
            process_one_pdf(str(p), str(output_folder), excel_path, existing_doc_keys)
        except Exception as e:
            print(f"[ERROR] Failed on {p.name}: {e}")
            continue


# ----------------------------
# Entrypoint
# ----------------------------
if __name__ == "__main__":
    input_folder = get_path("GIOIA_FIRST_ORDER_INPUT_FOLDER")
    output_folder = get_path("GIOIA_FIRST_ORDER_OUTPUT_FOLDER")
    if input_folder is None or output_folder is None:
        raise RuntimeError(
            "Set GIOIA_FIRST_ORDER_INPUT_FOLDER and GIOIA_FIRST_ORDER_OUTPUT_FOLDER in .env."
        )
    process_folder(str(input_folder), str(output_folder))
